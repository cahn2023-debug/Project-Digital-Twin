from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook

from app.importer import (
    CameraWorkbookProfile,
    ProfileStore,
    WorkbookProfile,
    parse_camera_rows,
    parse_camera_workbook,
    scan_workbook,
)


def test_camera_import_normalizes_values_and_retains_provenance() -> None:
    file_id = uuid4()
    cameras, issues = parse_camera_rows(
        [{
            "CameraCode": " CAM-001 ",
            "Name": " Main Camera ",
            "IP Address": "192.168.001.010",
            "Latitude": "13.7563",
            "Longitude": "100.5018",
        }],
        file_id=file_id,
        file_revision=7,
    )

    assert not issues
    assert cameras[0].code == "CAM-001"
    assert cameras[0].ip_address == "192.168.1.10"
    assert cameras[0].source is not None
    assert cameras[0].source.file_id == file_id
    assert cameras[0].source.file_revision == 7
    assert cameras[0].source.row == 2


def test_camera_import_reports_duplicate_and_invalid_rows() -> None:
    cameras, issues = parse_camera_rows([
        {"CameraCode": "CAM-001", "IP": "not-an-ip"},
        {"CameraCode": "CAM-001", "IP": "10.0.0.1"},
        {"Name": "missing code"},
    ])

    assert not cameras
    assert {issue.code for issue in issues} == {"INVALID_VALUE", "DUPLICATE_CAMERA_CODE", "MISSING_CAMERA_CODE"}


def test_workbook_scan_skips_hidden_sheets_discovers_regions_and_keeps_formula_source(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CAMERA"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "Camera"
    sheet.append(["CameraCode", "Name", "Status"])
    sheet.append(["CAM-001", "Main", "=1+1"])
    sheet.append(["TOTAL", None, None])
    sheet.append([None, None, None])
    sheet.append(["Other table", "Value", None])
    hidden = workbook.create_sheet("HIDDEN")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "Should not be scanned"

    path = tmp_path / "test-importer-scan.xlsx"
    workbook.save(path)
    result = scan_workbook(path, file_id=uuid4(), file_revision=3)

    assert result.skipped_sheets == ["HIDDEN"]
    assert len(result.regions) == 2
    camera_cells = {(cell.row, cell.column): cell for cell in result.cells if cell.sheet == "CAMERA"}
    assert camera_cells[(1, "B")].value == "Camera"
    assert camera_cells[(3, "C")].value == "=1+1"
    assert camera_cells[(3, "C")].formula == "=1+1"


def test_profile_parses_merged_headers_skips_totals_and_is_immutable(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CAMERA"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "Camera"
    sheet.append(["CameraCode", "Name", "Status"])
    sheet.append(["CAM-001", "Main", "DESIGNED"])
    sheet.append(["TOTAL", None, None])
    path = tmp_path / "profiled.xlsx"
    workbook.save(path)

    aliases = dict(CameraWorkbookProfile.aliases)
    aliases.update(
        {
            "code": ("Camera / CameraCode",),
            "name": ("Camera / Name",),
            "status": ("Camera / Status",),
        }
    )
    profile = WorkbookProfile(
        profile_id="camera",
        version=1,
        sheet="CAMERA",
        header_rows=(1, 2),
        data_start_row=3,
        table_start_row=1,
        skip_row_patterns=("total",),
        aliases=aliases,
    )
    cameras, issues = parse_camera_workbook(path, file_revision=4, profile=profile)

    assert not issues
    assert [camera.code for camera in cameras] == ["CAM-001"]
    store = ProfileStore(tmp_path / "profiles")
    saved = store.save(profile)
    assert saved.name == "camera.v1.json"
    assert store.load("camera", 1).header_rows == (1, 2)
    with pytest.raises(FileExistsError):
        store.save(profile)
