from hashlib import sha256

import pytest
from openpyxl import Workbook, load_workbook

from app.importer import parse_camera_workbook
from app.workbook import FileConflictError, file_sha256, write_camera_workbook


def test_managed_workbook_round_trip_preserves_unmanaged_sheet_and_column(tmp_path) -> None:
    path = tmp_path / "camera.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CAMERA"
    sheet.append(["CameraCode", "Name", "Status", "UnmanagedNote"])
    sheet.append(["CAM-001", "Main", "DESIGNED", "keep me"])
    notes = workbook.create_sheet("NOTES")
    notes["A1"] = "Do not remove"
    workbook.save(path)
    original_hash = file_sha256(path)

    updated_hash = write_camera_workbook(
        path,
        expected_sha256=original_hash,
        updates={"CAM-001": {"name": "Main Updated", "status": "AS_BUILT"}},
    )

    cameras, issues = parse_camera_workbook(path, file_revision=2)
    assert not issues
    assert cameras[0].name == "Main Updated"
    assert cameras[0].status == "AS_BUILT"
    assert updated_hash != original_hash
    rewritten = load_workbook(path, data_only=False)
    assert rewritten["CAMERA"]["D2"].value == "keep me"
    assert rewritten["NOTES"]["A1"].value == "Do not remove"
    assert (tmp_path / "camera.xlsx.bak").exists()


def test_managed_workbook_rejects_stale_hash_without_mutation(tmp_path) -> None:
    path = tmp_path / "camera.xlsx"
    workbook = Workbook()
    workbook.active.title = "CAMERA"
    workbook.active.append(["CameraCode", "Name"])
    workbook.active.append(["CAM-001", "Main"])
    workbook.save(path)
    original_bytes = path.read_bytes()

    with pytest.raises(FileConflictError):
        write_camera_workbook(path, expected_sha256=sha256(b"stale").hexdigest(), updates={"CAM-001": {"name": "Changed"}})

    assert path.read_bytes() == original_bytes
