from __future__ import annotations

import hashlib
import os
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .importer import CameraWorkbookProfile


class FileConflictError(Exception):
    pass


def file_sha256(path: str | Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def write_camera_workbook(
    path: str | Path,
    *,
    expected_sha256: str,
    updates: dict[str, dict[str, Any]],
) -> str:
    """Write only managed Camera columns after verifying the expected file version."""
    source_path = Path(path)
    actual_sha256 = file_sha256(source_path)
    if actual_sha256 != expected_sha256:
        raise FileConflictError(f"FILE_CONFLICT: expected {expected_sha256}, found {actual_sha256}")

    workbook = load_workbook(source_path, data_only=False)
    if CameraWorkbookProfile.sheet not in workbook.sheetnames:
        raise ValueError(f"Missing sheet {CameraWorkbookProfile.sheet}")
    sheet = workbook[CameraWorkbookProfile.sheet]
    headers = {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value is not None}

    def find_header(field_name: str) -> int | None:
        aliases = CameraWorkbookProfile.aliases[field_name]
        return next((headers[alias] for alias in aliases if alias in headers), None)

    code_column = find_header("code")
    if code_column is None:
        raise ValueError("Camera workbook has no managed Camera code column")
    managed_columns = {
        field_name: find_header(field_name)
        for field_name in CameraWorkbookProfile.aliases
        if field_name != "code"
    }
    found_codes: set[str] = set()
    for row in sheet.iter_rows(min_row=2):
        code_value = row[code_column - 1].value
        code = str(code_value).strip() if code_value is not None else ""
        if code not in updates:
            continue
        found_codes.add(code)
        for field_name, value in updates[code].items():
            column = managed_columns.get(field_name)
            if column is None:
                raise ValueError(f"Unsupported managed field: {field_name}")
            row[column - 1].value = value

    missing_codes = set(updates) - found_codes
    if missing_codes:
        raise ValueError(f"Camera codes not found: {sorted(missing_codes)}")

    backup_path = source_path.with_name(source_path.name + ".bak")
    temporary_path = source_path.with_name(source_path.name + ".project-digital-twin.tmp.xlsx")
    shutil.copy2(source_path, backup_path)
    try:
        workbook.save(temporary_path)
        validation_workbook = load_workbook(temporary_path, read_only=True, data_only=False)
        if CameraWorkbookProfile.sheet not in validation_workbook.sheetnames:
            raise ValueError("Written workbook failed profile validation")
        validation_workbook.close()
        os.replace(temporary_path, source_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()

    return file_sha256(source_path)
