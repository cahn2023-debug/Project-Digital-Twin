from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable
from uuid import UUID, uuid4

from openpyxl import load_workbook

from .domain import (
    Camera,
    ImportIssue,
    ImportResult,
    SourceLocator,
    normalize_coordinate,
    normalize_ip,
    normalize_text,
)


class CameraWorkbookProfile:
    sheet = "CAMERA"
    aliases = {
        "code": ("CameraCode", "Camera ID", "Mã Camera", "Code"),
        "name": ("Name", "Camera Name", "Tên Camera"),
        "intersection_id": ("IntersectionCode", "Intersection ID", "Intersection"),
        "manufacturer": ("Manufacturer", "Hãng"),
        "model": ("Model",),
        "ip_address": ("IP", "IP Address", "IPAddress"),
        "status": ("Status", "Trạng thái"),
        "latitude": ("Latitude", "Lat", "Vĩ độ"),
        "longitude": ("Longitude", "Lng", "Lon", "Kinh độ"),
    }


def _header_map(headers: Iterable[Any]) -> dict[str, str]:
    return {str(value).strip(): str(value).strip() for value in headers if value is not None}


def _find_column(headers: dict[str, str], aliases: tuple[str, ...]) -> str | None:
    for alias in aliases:
        if alias in headers:
            return headers[alias]
    return None


def parse_camera_rows(
    rows: Iterable[dict[str, Any]],
    *,
    file_id: UUID | None = None,
    file_revision: int = 1,
    sheet: str = CameraWorkbookProfile.sheet,
) -> tuple[list[Camera], list[ImportIssue]]:
    file_id = file_id or uuid4()
    normalized: list[Camera] = []
    issues: list[ImportIssue] = []
    seen_codes: set[str] = set()

    for row_number, raw in enumerate(rows, start=2):
        headers = _header_map(raw.keys())
        code_column = _find_column(headers, CameraWorkbookProfile.aliases["code"])
        code = normalize_text(raw.get(code_column)) if code_column else None
        if code is None:
            issues.append(ImportIssue("MISSING_CAMERA_CODE", "Camera code is required", row_number, code_column))
            continue
        if code in seen_codes:
            issues.append(ImportIssue("DUPLICATE_CAMERA_CODE", f"Duplicate Camera code: {code}", row_number, code_column))
            continue
        seen_codes.add(code)

        try:
            ip_column = _find_column(headers, CameraWorkbookProfile.aliases["ip_address"])
            latitude_column = _find_column(headers, CameraWorkbookProfile.aliases["latitude"])
            longitude_column = _find_column(headers, CameraWorkbookProfile.aliases["longitude"])
            latitude = normalize_coordinate(raw.get(latitude_column), -90, 90, "latitude") if latitude_column else None
            longitude = normalize_coordinate(raw.get(longitude_column), -180, 180, "longitude") if longitude_column else None
            ip = normalize_ip(raw.get(ip_column)) if ip_column else None
        except ValueError as exc:
            issues.append(ImportIssue("INVALID_VALUE", str(exc), row_number))
            continue

        def value(field_name: str) -> str | None:
            column = _find_column(headers, CameraWorkbookProfile.aliases[field_name])
            return normalize_text(raw.get(column)) if column else None

        locator = SourceLocator(
            file_id=file_id,
            file_revision=file_revision,
            sheet=sheet,
            row=row_number,
            column=code_column or "",
        )
        normalized.append(
            Camera(
                entity_id=uuid4(),
                project_id=UUID(int=0),
                code=code,
                name=value("name"),
                intersection_id=None,
                manufacturer=value("manufacturer"),
                model=value("model"),
                ip_address=ip,
                status=value("status"),
                properties={
                    "latitude": latitude,
                    "longitude": longitude,
                    "raw": dict(raw),
                },
                source=locator,
            )
        )
    return normalized, issues


def parse_camera_workbook(
    path: str | Path,
    *,
    file_id: UUID | None = None,
    file_revision: int = 1,
) -> tuple[list[Camera], list[ImportIssue]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    if CameraWorkbookProfile.sheet not in workbook.sheetnames:
        return [], [ImportIssue("UNMAPPED_SHEET", f"Missing sheet {CameraWorkbookProfile.sheet}", 0)]
    sheet = workbook[CameraWorkbookProfile.sheet]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return [], [ImportIssue("UNSUPPORTED_ROW", "Workbook sheet is empty", 0)]
    dictionaries = [dict(zip(headers, row)) for row in rows]
    return parse_camera_rows(dictionaries, file_id=file_id, file_revision=file_revision)
