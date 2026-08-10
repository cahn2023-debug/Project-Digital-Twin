from __future__ import annotations

import hashlib
import os
import shutil
from tempfile import NamedTemporaryFile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..importers.excel import CameraWorkbookProfile


class FileConflictError(Exception):
    pass


class FileLockedError(Exception):
    pass


class WriteConfirmationError(Exception):
    pass


def file_sha256(path: str | Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def write_camera_workbook(
    path: str | Path,
    *,
    expected_sha256: str,
    updates: dict[str, dict[str, Any]],
    confirmed: bool = False,
    header_row: int = 1,
    data_start_row: int = 2,
    backup_path: str | Path | None = None,
) -> str:
    """Write confirmed managed fields while preserving the rest of the workbook."""
    if not confirmed:
        raise WriteConfirmationError("WRITE_CONFIRMATION_REQUIRED")
    source_path = Path(path)
    actual_sha256 = file_sha256(source_path)
    if actual_sha256 != expected_sha256:
        raise FileConflictError(f"FILE_CONFLICT: expected {expected_sha256}, found {actual_sha256}")

    try:
        with source_path.open("r+b"):
            pass
    except PermissionError as exc:
        raise FileLockedError(f"FILE_LOCKED: {source_path}") from exc

    keep_vba = source_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(source_path, data_only=False, keep_vba=keep_vba)
    if CameraWorkbookProfile.sheet not in workbook.sheetnames:
        raise ValueError(f"Missing sheet {CameraWorkbookProfile.sheet}")
    sheet = workbook[CameraWorkbookProfile.sheet]
    headers = {
        " ".join(str(cell.value).strip().split()).casefold(): cell.column
        for cell in sheet[header_row]
        if cell.value is not None
    }

    def find_header(field_name: str) -> int | None:
        aliases = CameraWorkbookProfile.aliases[field_name]
        return next((headers[" ".join(alias.split()).casefold()] for alias in aliases if " ".join(alias.split()).casefold() in headers), None)

    code_column = find_header("code")
    if code_column is None:
        raise ValueError("Camera workbook has no managed Camera code column")
    managed_columns: dict[str, int] = {}
    for field_name in dict.fromkeys(field_name for update in updates.values() for field_name in update):
        if field_name == "code":
            continue
        column = find_header(field_name) if field_name in CameraWorkbookProfile.aliases else None
        if column is None:
            column = sheet.max_column + 1
            header = CameraWorkbookProfile.aliases.get(field_name, (field_name,))[0]
            sheet.cell(header_row, column).value = header
            headers[" ".join(header.split()).casefold()] = column
        managed_columns[field_name] = column

    found_codes: set[str] = set()
    for row in sheet.iter_rows(min_row=data_start_row):
        code_value = row[code_column - 1].value
        code = str(code_value).strip() if code_value is not None else ""
        if code not in updates:
            continue
        found_codes.add(code)
        for field_name, value in updates[code].items():
            column = managed_columns.get(field_name)
            if column is None:
                raise ValueError(f"Unsupported managed field: {field_name}")
            row[column - 1].value = value

    missing_codes = set(updates) - found_codes
    if missing_codes:
        raise ValueError(f"Camera codes not found: {sorted(missing_codes)}")

    backup = Path(backup_path) if backup_path is not None else source_path.with_name(source_path.name + ".bak")
    temporary_path: Path | None = None
    try:
        try:
            shutil.copy2(source_path, backup)
        except PermissionError as exc:
            raise FileLockedError(f"FILE_LOCKED: {source_path}") from exc
        with NamedTemporaryFile(
            dir=source_path.parent,
            prefix=f".{source_path.name}.",
            suffix=".project-digital-twin.tmp.xlsx",
            delete=False,
        ) as temporary:
            temporary_path = Path(temporary.name)
        workbook.save(temporary_path)
        validation_workbook = load_workbook(temporary_path, read_only=True, data_only=False, keep_vba=keep_vba)
        if CameraWorkbookProfile.sheet not in validation_workbook.sheetnames:
            raise ValueError("Written workbook failed profile validation")
        validation_workbook.close()
        try:
            os.replace(temporary_path, source_path)
        except PermissionError as exc:
            raise FileLockedError(f"FILE_LOCKED: {source_path}") from exc
    finally:
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()

    return file_sha256(source_path)
