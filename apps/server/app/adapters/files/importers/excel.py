from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable, Mapping
from uuid import UUID, uuid4

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ....domain import (
    Camera,
    ImportIssue,
    ImportResult,
    SourceLocator,
    normalize_coordinate,
    normalize_ip,
    normalize_text,
)


class CameraWorkbookProfile:
    sheet = "CAMERA"
    aliases = {
        "code": ("CameraCode", "Camera ID", "Mã Camera", "Code"),
        "name": ("Name", "Camera Name", "Tên Camera"),
        "intersection_id": ("IntersectionCode", "Intersection ID", "Intersection"),
        "manufacturer": ("Manufacturer", "Hãng"),
        "model": ("Model",),
        "ip_address": ("IP", "IP Address", "IPAddress"),
        "status": ("Status", "Trạng thái"),
        "latitude": ("Latitude", "Lat", "Vĩ độ"),
        "longitude": ("Longitude", "Lng", "Lon", "Kinh độ"),
    }


@dataclass(frozen=True)
class WorkbookProfile:
    profile_id: str
    version: int
    sheet: str
    header_rows: tuple[int, ...] = (1,)
    data_start_row: int = 2
    table_start_row: int | None = None
    skip_row_patterns: tuple[str, ...] = ()
    aliases: Mapping[str, tuple[str, ...]] = field(default_factory=lambda: dict(CameraWorkbookProfile.aliases))


@dataclass(frozen=True)
class WorkbookCell:
    file_id: UUID
    file_revision: int
    sheet: str
    row: int
    column: str
    value: Any
    formula: str | None = None


@dataclass(frozen=True)
class TableRegion:
    sheet: str
    start_row: int
    end_row: int
    min_column: int
    max_column: int
    header_candidates: tuple[int, ...]


@dataclass
class WorkbookScanResult:
    regions: list[TableRegion] = field(default_factory=list)
    cells: list[WorkbookCell] = field(default_factory=list)
    skipped_sheets: list[str] = field(default_factory=list)


class ProfileStore:
    """Small immutable JSON profile store used by the parser boundary."""

    def __init__(self, root: str | Path) -> None:
        self.root = Path(root)
        self.root.mkdir(parents=True, exist_ok=True)

    def save(self, profile: WorkbookProfile) -> Path:
        path = self.root / f"{profile.profile_id}.v{profile.version}.json"
        if path.exists():
            raise FileExistsError(f"Profile version already exists: {path.name}")
        temporary = path.with_suffix(".tmp")
        temporary.write_text(json.dumps(asdict(profile), ensure_ascii=False, indent=2), encoding="utf-8")
        temporary.replace(path)
        return path

    def load(self, profile_id: str, version: int) -> WorkbookProfile:
        path = self.root / f"{profile_id}.v{version}.json"
        payload = json.loads(path.read_text(encoding="utf-8"))
        payload["header_rows"] = tuple(payload["header_rows"])
        payload["skip_row_patterns"] = tuple(payload["skip_row_patterns"])
        payload["aliases"] = {key: tuple(value) for key, value in payload["aliases"].items()}
        return WorkbookProfile(**payload)


def _merged_anchor(sheet: Any, row: int, column: int) -> tuple[int, int]:
    for merged_range in sheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return merged_range.min_row, merged_range.min_col
    return row, column


def _cell_value(formula_sheet: Any, value_sheet: Any, row: int, column: int) -> tuple[Any, str | None]:
    anchor_row, anchor_column = _merged_anchor(formula_sheet, row, column)
    formula_value = formula_sheet.cell(anchor_row, anchor_column).value
    cached_value = value_sheet.cell(anchor_row, anchor_column).value
    if isinstance(formula_value, str) and formula_value.startswith("="):
        return (cached_value if cached_value is not None else formula_value), formula_value
    return formula_value, None


def scan_workbook(
    path: str | Path,
    *,
    file_id: UUID | None = None,
    file_revision: int = 1,
) -> WorkbookScanResult:
    file_id = file_id or uuid4()
    formula_workbook = load_workbook(path, read_only=False, data_only=False)
    value_workbook = load_workbook(path, read_only=False, data_only=True)
    result = WorkbookScanResult()

    for sheet_name in formula_workbook.sheetnames:
        formula_sheet = formula_workbook[sheet_name]
        if formula_sheet.sheet_state != "visible":
            result.skipped_sheets.append(sheet_name)
            continue
        value_sheet = value_workbook[sheet_name]
        row_columns: dict[int, list[int]] = {}
        for row in range(1, formula_sheet.max_row + 1):
            for column in range(1, formula_sheet.max_column + 1):
                value, formula = _cell_value(formula_sheet, value_sheet, row, column)
                if value is None or value == "":
                    continue
                column_name = get_column_letter(column)
                result.cells.append(
                    WorkbookCell(
                        file_id=file_id,
                        file_revision=file_revision,
                        sheet=sheet_name,
                        row=row,
                        column=column_name,
                        value=value,
                        formula=formula,
                    )
                )
                row_columns.setdefault(row, []).append(column)

        occupied_rows = sorted(row_columns)
        if not occupied_rows:
            continue
        start = previous = occupied_rows[0]
        for current in occupied_rows[1:] + [None]:
            if current is not None and current == previous + 1:
                previous = current
                continue
            columns = [column for row in range(start, previous + 1) for column in row_columns[row]]
            result.regions.append(
                TableRegion(
                    sheet=sheet_name,
                    start_row=start,
                    end_row=previous,
                    min_column=min(columns),
                    max_column=max(columns),
                    header_candidates=tuple(range(start, min(previous, start + 2) + 1)),
                )
            )
            if current is not None:
                start = previous = current
    return result


def _region_for_profile(scan: WorkbookScanResult, profile: WorkbookProfile) -> TableRegion | None:
    candidates = [region for region in scan.regions if region.sheet == profile.sheet]
    if profile.table_start_row is not None:
        candidates = [region for region in candidates if region.start_row == profile.table_start_row]
    return candidates[0] if candidates else None


def _profile_rows(scan: WorkbookScanResult, profile: WorkbookProfile) -> list[dict[str, Any]]:
    region = _region_for_profile(scan, profile)
    if region is None:
        return []
    cells = {
        (cell.row, cell.column): cell
        for cell in scan.cells
        if cell.sheet == profile.sheet
    }
    headers: list[str] = []
    for column_number in range(region.min_column, region.max_column + 1):
        column_name = get_column_letter(column_number)
        parts: list[str] = []
        for header_row in profile.header_rows:
            cell = cells.get((header_row, column_name))
            if cell is not None and str(cell.value).strip() and str(cell.value).strip() not in parts:
                parts.append(str(cell.value).strip())
        headers.append(" / ".join(parts) or column_name)

    rows: list[dict[str, Any]] = []
    for row_number in range(profile.data_start_row, region.end_row + 1):
        values = {
            header: cells[(row_number, get_column_letter(region.min_column + index))].value
            for index, header in enumerate(headers)
            if (row_number, get_column_letter(region.min_column + index)) in cells
        }
        if not values:
            continue
        joined = " ".join(str(value) for value in values.values() if value is not None).casefold()
        if any(pattern.casefold() in joined for pattern in profile.skip_row_patterns):
            continue
        rows.append(values)
    return rows


def _header_map(headers: Iterable[Any]) -> dict[str, str]:
    return {str(value).strip(): str(value).strip() for value in headers if value is not None}


def _find_column(headers: dict[str, str], aliases: tuple[str, ...]) -> str | None:
    for alias in aliases:
        if alias in headers:
            return headers[alias]
    return None


def parse_camera_rows(
    rows: Iterable[dict[str, Any]],
    *,
    file_id: UUID | None = None,
    file_revision: int = 1,
    sheet: str = CameraWorkbookProfile.sheet,
    row_start: int = 2,
    aliases: Mapping[str, tuple[str, ...]] | None = None,
) -> tuple[list[Camera], list[ImportIssue]]:
    file_id = file_id or uuid4()
    aliases = aliases or CameraWorkbookProfile.aliases
    normalized: list[Camera] = []
    issues: list[ImportIssue] = []
    seen_codes: set[str] = set()

    for row_number, raw in enumerate(rows, start=row_start):
        headers = _header_map(raw.keys())
        code_column = _find_column(headers, aliases["code"])
        code = normalize_text(raw.get(code_column)) if code_column else None
        if code is None:
            issues.append(ImportIssue("MISSING_CAMERA_CODE", "Camera code is required", row_number, code_column))
            continue
        if code in seen_codes:
            issues.append(ImportIssue("DUPLICATE_CAMERA_CODE", f"Duplicate Camera code: {code}", row_number, code_column))
            continue
        seen_codes.add(code)

        try:
            ip_column = _find_column(headers, aliases["ip_address"])
            latitude_column = _find_column(headers, aliases["latitude"])
            longitude_column = _find_column(headers, aliases["longitude"])
            latitude = normalize_coordinate(raw.get(latitude_column), -90, 90, "latitude") if latitude_column else None
            longitude = normalize_coordinate(raw.get(longitude_column), -180, 180, "longitude") if longitude_column else None
            ip = normalize_ip(raw.get(ip_column)) if ip_column else None
        except ValueError as exc:
            issues.append(ImportIssue("INVALID_VALUE", str(exc), row_number))
            continue

        def value(field_name: str) -> str | None:
            column = _find_column(headers, aliases[field_name])
            return normalize_text(raw.get(column)) if column else None

        locator = SourceLocator(
            file_id=file_id,
            file_revision=file_revision,
            sheet=sheet,
            row=row_number,
            column=code_column or "",
        )
        normalized.append(
            Camera(
                entity_id=uuid4(),
                project_id=UUID(int=0),
                code=code,
                name=value("name"),
                intersection_id=None,
                manufacturer=value("manufacturer"),
                model=value("model"),
                ip_address=ip,
                status=value("status"),
                properties={
                    "latitude": latitude,
                    "longitude": longitude,
                    "raw": dict(raw),
                },
                source=locator,
            )
        )
    return normalized, issues


def parse_camera_workbook(
    path: str | Path,
    *,
    file_id: UUID | None = None,
    file_revision: int = 1,
    profile: WorkbookProfile | None = None,
) -> tuple[list[Camera], list[ImportIssue]]:
    _, cameras, issues, _ = parse_camera_workbook_for_import(
        path,
        file_id=file_id,
        file_revision=file_revision,
        profile=profile,
    )
    return cameras, issues


def parse_camera_workbook_for_import(
    path: str | Path,
    *,
    file_id: UUID | None = None,
    file_revision: int = 1,
    profile: WorkbookProfile | None = None,
) -> tuple[list[dict[str, Any]], list[Camera], list[ImportIssue], WorkbookScanResult]:
    profile = profile or WorkbookProfile(
        profile_id="camera-default",
        version=1,
        sheet=CameraWorkbookProfile.sheet,
        header_rows=(1,),
        data_start_row=2,
        table_start_row=1,
    )
    scan = scan_workbook(path, file_id=file_id, file_revision=file_revision)
    if profile.sheet not in {region.sheet for region in scan.regions}:
        return [], [], [ImportIssue("UNMAPPED_SHEET", f"Missing sheet {profile.sheet}", 0)], scan
    if _region_for_profile(scan, profile) is None:
        return [], [], [ImportIssue("UNMAPPED_TABLE", f"Missing table in sheet {profile.sheet}", 0)], scan
    rows = _profile_rows(scan, profile)
    if not rows:
        return [], [], [ImportIssue("UNSUPPORTED_ROW", "Workbook table has no data rows", 0)], scan
    cameras, issues = parse_camera_rows(
        rows,
        file_id=file_id,
        file_revision=file_revision,
        sheet=profile.sheet,
        row_start=profile.data_start_row,
        aliases=profile.aliases,
    )
    return rows, cameras, issues, scan
